#! /library/Frameworks/Python.framework/Versions/3.5/python3.5
# In this tutorial we will be working with openpyxl module
# ===================================================================================================
# ______________________ Openpyxl module ______________________
# --first you import the module
# --then you use load_workbook() function
# wb = openpyxl.load_workbook('name') - load the workbook by the name
# ______________________ Working with workbook and sheets ______________________
# wb.get_sheet_names() - return the list of sheet names
# sheet = wb.get_sheet_by_name() - return certain sheet to the variable
# sheet.title - return the title of the sheet
# sheet = wb.get_active_sheet() - return active sheet into a variable
# sheet['cell name'] - return the object of a cell into a variable
# sheet['cell name'].value - return the value of the specified cell
# sheet['cell name'].row - return the row of specified cell
# sheet['cell name'].column - return column of the specified cell
# sheet['cell name'].coordinate - return the coordinate of the cell (A1, B100, etc.)
# cell(row=num, column=num) - method to locate cell by giving numbered positions
# --you can also determine the sheet size by highest row and highest cell
# sheet.max_row - returns the number of the highest row
# sheet.max_column - returns the number of the highest column
# --you can also convert index to letter and letter to index
# --for that you need to import two functios as below
# get_column_letter('num') - returns the letter from the index
# get_column_index_from_string('letter') - returns the index from the letter
# -- you can also slice the sheet in the way you want
# slice = sheet['A1': 'C3'] - slice the sheet into a section defined by A1 to C3
#                           - slice will be a set touples for each row
#                           - working through slice you work through the rows first, then you go over cells
# --you can also access certain rows and columns
# sheet.columns[1] - access the column by index
# sheet.rows[1] - access the row by index
# ______________________ Writing in excel documents ______________________
# --create a new blank workbook object
# wb = openpyxl.Workbook()
# --modify things
# --save the file
# wb.save('directory') - save the existing wb object into the folder
# wb.create_sheet() - create a new sheet in a workbook
#                   - you can specify the index=0 for the sheet to be the first one
#                   - wb.create_sheet(index=0, title='Hello')
# --to remove the sheet you have to pass an object, not a sheet name
# wb.remove_sheet(wb.get_sheet_by_name('name')) - remove the sheet
# sheet['A1'] = 55 - change the value of the cell
# ______________________ Working with fonts and styles ______________________
# --import the necessary components first
# from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# --then define your parameters, for example here we will define font
# font_obj = Font(italic=True, etc) - create Font object
#                                   - name=Calibri
#                                   - size=55
#                                   - bold=True
#                                   - italic=True
# --apply font style
# sheet['A1'].font = font_obj
# --you can do the same thing for other parameters, such as Border, Alignment, etc.
# ______________________ Formulas ______________________
# sheet['B9'] = '=SUM(B1:B9)' - creating the sum formula
# ______________________ Adjusting rows and columns ______________________
# sheet.row_dimensions[1].height = 70 - setting up the dimensions of the row 1
# sheet.column_dimensions['B'].width = 60 - setting the dimensions for the column
# ______________________ Merging and un-merging cells ______________________
# sheet.merge_cells('A1:D3') - merge a range of cells
# sheet.unmerge_cells('A1:D3') - un-merge the range of cells
# ______________________ Freezing panes ______________________
# sheet.freeze_panes = 'A2' - row 1
# sheet.freeze_panes = 'B1' - column A
# sheet.freeze_panes = 'C1' - column A and B
# sheet.freeze_panes = 'C2' - row 1 and columns A and B
# sheet.freeze_panes = 'A1' - unfreeze panes
# ______________________ Charts ______________________

# ===================================================================================================

# loading the module and opening the excel file
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import BarChart, Reference, Series
wb = openpyxl.load_workbook('1_Working_files/example.xlsx')
print(type(wb))

# loading the worksheet into a variable
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.title)

# let's now work with cells in the excel sheet
print(sheet['a1'].value, ' ', sheet['b1'].value, ' ', sheet['c1'].value)
print(sheet['a1'].column)
print(sheet['a1'].coordinate)

# you can also use cell() method to locate the cell
print(sheet.cell(row=100, column=200).coordinate, '\n')
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# Let's determine the highest rows and columns in the sheet1
print('highest row is: ', sheet.max_row)
print('highest column is: ', sheet.max_column)

# lets convert index to letter and letter to index
print(get_column_letter(111))
print(column_index_from_string('DG'))

# let's slice the sheet
# for first retrieves the row, then we go through each item in the row
slice1 = sheet['A1': 'C3']
for i in slice1:
    for j in i:
        print(j.value, end=', ')
    print('\n')

# let's access certain columns and rows
print(sheet.columns[1])
print(sheet.rows[1])
for i in sheet.columns[1]:
    if i.value is not None:
        print(i.value)

# let's open the new blank workbook
wb = openpyxl.Workbook()
print(wb.get_sheet_names())
sheet_new = wb.get_sheet_by_name('Sheet')
sheet_new.title = 'My title'
print(sheet_new.title)
wb.save('1_Working_files/playful.xlsx')
wb.create_sheet('Hello')
print(wb.get_sheet_names())

# let's open the new workbook and play with some fomratting
wb1 = openpyxl.Workbook()
sheet = wb1.get_sheet_by_name('Sheet')
font_obj = Font(size=24, italic=True)
sheet['A1'].font = font_obj
sheet['A1'].value = 'Hey'
wb1.save('1_Working_files/fonts_styles.xlsx')

# Let's create a chart in the excel file
wb_chart = openpyxl.Workbook()
sheet_chart = wb_chart.get_sheet_by_name('Sheet')
for i in range(1, 11):
    sheet_chart['A' + str(i)].value = i
values = openpyxl.chart.Reference(sheet_chart, min_col=1, min_row=1, max_col=1, max_row=10)
chart = openpyxl.chart.BarChart()
chart.add_data(values)
sheet_chart.add_chart(chart)
wb_chart.save('1_Working_files/example_chart.xlsx')





